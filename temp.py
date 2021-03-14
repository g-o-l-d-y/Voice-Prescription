# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""


import openpyxl 
import pymongo
  

wrkbk = openpyxl.load_workbook("C:/Users/goldy/.spyder-py3/SIH 2020/final.xlsx") 
  
sh = wrkbk.active 
  
for row in sh.iter_rows(min_row=1, min_col=1, max_row=100, max_col=3): 
    for cell in row: 
        print(cell.value, end=" ") 
    print() 

myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["Prescriptions"]
mycol = mydb["drugs"]


dic={}

for row in sh.iter_rows(): 
    l=[]
    for cell in row:
        l+=[cell.value]
    if l[1] not in dic:
        dic[l[1].lower()]=[]
        dic[l[1].lower()].append(l[2])
    else:
        if l[2] not in dic[l[1].lower()]:
            dic[l[1].lower()].append(l[2])
for k,v in dic.items():
    mycol.insert_one({"name":k,"medicine":v})
            